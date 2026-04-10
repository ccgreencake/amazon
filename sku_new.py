import os
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import date

# --- 配置区域 ---
folder_path = "F:\\上架\\260331"
# 模式选择：
# 'exclude_merge' -> 排除文件名带 "merge" 的文件
# 'only_merge'    -> 只汇总文件名带 "merge" 的文件
filter_mode = 'exclude_merge'

# 获取今天的日期
today = date.today()
date_str = today.strftime("%Y%m%d")

folder_name = os.path.basename(folder_path)
print(f"正在处理文件夹: {folder_name}")

# 新建 .xlsx 文件路径
xlsx_file_path = f"C:\\Users\\Administrator\\Desktop\\upload_sku\\sku{folder_name}.xlsx"

# 创建新的 Workbook 对象
wb_dest = Workbook()
ws_dest = wb_dest.active
ws_dest.title = 'Merged Data'

# 添加标题行
ws_dest["A1"] = "系统SKU"
ws_dest["B1"] = "平台SKU"

original_data = []
processed_data = []

# --- 遍历文件夹 ---
for filename in os.listdir(folder_path):
    # 1. 跳过临时文件
    if filename.startswith('~$'):
        continue

    # 2. 检查是否是 Excel 文件
    if not filename.endswith('.xlsx'):
        continue

    # 3. 根据变量进行文件名过滤逻辑
    is_merge_file = "merge" in filename.lower()  # 检查文件名是否包含 merge (不区分大小写)

    if filter_mode == 'exclude_merge':
        if is_merge_file:
            print(f"跳过文件 (排除模式): {filename}")
            continue
    elif filter_mode == 'only_merge':
        if not is_merge_file:
            print(f"跳过文件 (仅匹配模式): {filename}")
            continue

    # 拼接文件的完整路径
    file_path = os.path.join(folder_path, filename)
    print(f"正在读取: {filename}")

    # 加载文件
    try:
        # 使用 data_only=True 可以读取公式计算后的值（如果有的话）
        wb_src = load_workbook(filename=file_path, read_only=True, data_only=True)

        # --- 修改后的工作表选择逻辑 ---
        ws_name = None
        if 'Template' in wb_src.sheetnames:
            ws_name = 'Template'
        elif 'Vorlage' in wb_src.sheetnames:
            ws_name = 'Vorlage'

        if ws_name is None:
            print(f"警告: 文件 {filename} 中既找不到 'Template' 也找不到 'Vorlage'")
            wb_src.close()
            continue

        ws_src = wb_src[ws_name]
        print(f"成功找到工作表: [{ws_name}] 在文件: {filename}")

        # 遍历读取数据
        for row in ws_src.iter_rows(min_row=4, max_col=2, values_only=True):
            if row and len(row) >= 2 and row[1] is not None:
                sku = str(row[1]).strip()
                original_data.append(sku)

                # SKU 清洗逻辑
                if sku.startswith(('V', 'T', 'C', 'Z', 'P')):
                    sku = sku[20:]
                elif sku.startswith(('li', 'yo', 'lc')):
                    sku = sku[10:]
                processed_data.append(sku)

        wb_src.close()
    except Exception as e:
        print(f"处理文件 {filename} 时出错: {e}")

# --- 写入数据到新表 ---
# 将清洗后的数据保存到第一列 (系统SKU)
for i, sku in enumerate(processed_data):
    ws_dest.cell(row=i + 2, column=1, value=sku)

# 将原始数据保存到第二列 (平台SKU)
for i, sku in enumerate(original_data):
    ws_dest.cell(row=i + 2, column=2, value=sku)

# 保存文件
try:
    # 检查保存目录是否存在，不存在则创建
    os.makedirs(os.path.dirname(xlsx_file_path), exist_ok=True)
    wb_dest.save(filename=xlsx_file_path)
    print(f"处理完成！文件已保存至: {xlsx_file_path}")
except Exception as e:
    print(f"保存文件失败: {e}")

wb_dest.close()