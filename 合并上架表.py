from datetime import datetime
import os
import shutil
import subprocess  # 导入用于打开程序的模块
from openpyxl import load_workbook


def merge_xlsx(folder_path, header_rows=3, use_grouping=True, exclude_merge=True):
    """
    :param exclude_merge: 如果为 True, 排除文件名中包含 'merge' 的文件
    """
    # 1. 获取文件夹内所有有效的xlsx文件
    raw_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx') and not f.startswith('~$')]

    all_files = []
    for f in raw_files:
        if exclude_merge and 'merge' in f.lower():
            continue
        all_files.append(f)

    if not all_files:
        print(f"文件夹内没有找到可处理的 xlsx 文件。")
        return

    folder_name = os.path.basename(os.path.abspath(folder_path))

    # 2. 分组逻辑
    groups = {}
    if use_grouping:
        for f in all_files:
            prefix = f[:2]
            groups.setdefault(prefix, []).append(f)
    else:
        groups[''] = all_files

    # 3. 执行合并逻辑
    for prefix, files in groups.items():
        if not files:
            continue

        if use_grouping:
            output_filename = f"{prefix}{folder_name}merge.xlsx"
        else:
            output_filename = f"{folder_name}_merge.xlsx"

        output_path = os.path.join(folder_path, output_filename)

        if output_filename in files:
            files.remove(output_filename)
            if not files: continue

        if os.path.exists(output_path):
            try:
                os.remove(output_path)
            except Exception as e:
                print(f"无法删除已存在的输出文件 {output_filename}: {e}")
                continue

        print(f"\n>>> 正在生成: {output_filename}")

        # --- 核心修改点 1: 确定母版文件的目标 Sheet ---
        # 复制第一个文件作为母版
        shutil.copy(os.path.join(folder_path, files[0]), output_path)
        wb_target = load_workbook(output_path)

        # 检查母版里用哪个 Sheet
        if "Template" in wb_target.sheetnames:
            target_sheet_name = "Template"
        elif "Vorlage" in wb_target.sheetnames:
            target_sheet_name = "Vorlage"
        else:
            print(f"警告: 文件 {files[0]} 中既没有 'Template' 也没有 'Vorlage'，跳过该组。")
            wb_target.close()
            continue

        ws_target = wb_target[target_sheet_name]
        print(f"使用工作表: {target_sheet_name}")

        # 遍历剩余文件
        for i in range(1, len(files)):
            source_file_path = os.path.join(folder_path, files[i])
            wb_source = load_workbook(source_file_path, data_only=True)

            # --- 核心修改点 2: 检查源文件中的 Sheet ---
            current_source_sheet = None
            if "Template" in wb_source.sheetnames:
                current_source_sheet = wb_source["Template"]
            elif "Vorlage" in wb_source.sheetnames:
                current_source_sheet = wb_source["Vorlage"]

            if current_source_sheet is None:
                print(f"跳过文件 {files[i]}: 未找到 'Template' 或 'Vorlage'")
                wb_source.close()
                continue

            # 写入数据
            for row in current_source_sheet.iter_rows(min_row=header_rows + 1, values_only=True):
                if all(cell is None for cell in row):
                    continue
                ws_target.append(row)
            wb_source.close()

        # 保存结果
        wb_target.save(output_path)
        wb_target.close()
        print(f"--- {output_filename} 合并完成！---")

        # ==========================================
        # 自动打开功能
        # ==========================================
        if os.name == 'nt':
            wps_path = r'F:\APP\wps\WPS Office\ksolaunch.exe'
            if os.path.exists(wps_path):
                subprocess.Popen([wps_path, output_path])
                print(f"已调用WPS打开: {output_filename}")
            else:
                print(f"警告：找不到WPS路径 {wps_path}。")


# ==========================================
#               参数设置区域
# ==========================================
if __name__ == "__main__":
    # 自动获取当前日期
    folder_date = datetime.now().strftime('%y%m%d')
    # 设置 F 盘路径
    MY_FOLDER = os.path.join(r"F:\上架", folder_date)

    if not os.path.exists(MY_FOLDER):
        print(f"路径不存在: {MY_FOLDER}")
    else:
        merge_xlsx(
            folder_path=MY_FOLDER,
            header_rows=3,
            use_grouping=True,
            exclude_merge=True
        )