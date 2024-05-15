import openpyxl

# 读取包含数据的XLSX文件
input_file = r'C:\Users\123\Desktop\工作簿4.xlsx'  # 替换为您的输入文件路径
output_file = 'output.xlsx'  # 替换为您的输出文件路径

# 读取数据文件
workbook = openpyxl.load_workbook(input_file)
sheet = workbook.active

# 获取数据的列
column_data = [cell.value for cell in sheet['A']]


num_rows = 80
num_columns = -(-len(column_data) // num_rows)  # 向上取整，计算所需的列数
split_data = [column_data[i:i+num_rows] for i in range(0, len(column_data), num_rows)]

# 创建新的XLSX文件并写入拆分后的结果
output_workbook = openpyxl.Workbook()
output_sheet = output_workbook.active

for col_idx, column in enumerate(split_data, start=1):
    for row_idx, value in enumerate(column, start=1):
        output_sheet.cell(row=row_idx, column=col_idx, value=value)

# 保存新的XLSX文件
output_workbook.save(output_file)

print("拆分完成，结果已保存到", output_file)