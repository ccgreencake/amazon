from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
import os
import shutil
import tempfile
from pathlib import Path
from io import BytesIO
import re
from typing import List


class ExcelMerger:
    def __init__(self):
        self.merged_wb = Workbook()
        self.merged_ws = self.merged_wb.active
        self.current_row = 1
        self.header_copied = False
        self.column_widths = {}
        self.row_heights = {}
        self.images = []  # 存储图片信息字典

    def _copy_cell_style(self, source_cell, target_cell):
        """深度复制单元格所有样式属性"""
        try:
            # 字体
            target_cell.font = Font(
                name=source_cell.font.name,
                size=source_cell.font.size,
                bold=source_cell.font.bold,
                italic=source_cell.font.italic,
                underline=source_cell.font.underline,
                strike=source_cell.font.strike,
                color=source_cell.font.color
            )

            # 填充
            if source_cell.fill:
                target_cell.fill = PatternFill(
                    fill_type=source_cell.fill.fill_type,
                    start_color=source_cell.fill.start_color,
                    end_color=source_cell.fill.end_color
                )

            # 边框
            if source_cell.border:
                border = Border()
                for side in ['left', 'right', 'top', 'bottom']:
                    side_attr = getattr(source_cell.border, side)
                    if side_attr:
                        setattr(border, side, Side(
                            style=side_attr.style,
                            color=side_attr.color
                        ))
                target_cell.border = border

            # 对齐
            if source_cell.alignment:
                target_cell.alignment = Alignment(
                    horizontal=source_cell.alignment.horizontal,
                    vertical=source_cell.alignment.vertical,
                    wrap_text=source_cell.alignment.wrap_text,
                    shrink_to_fit=source_cell.alignment.shrink_to_fit,
                    indent=source_cell.alignment.indent,
                    text_rotation=source_cell.alignment.text_rotation
                )

            # 数字格式
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format

        except Exception as e:
            print(f"复制样式时出错: {e}")

    def _copy_row(self, source_ws, row_index, target_ws, target_row):
        """复制整行及其样式"""
        target_ws.row_dimensions[target_row].height = source_ws.row_dimensions[row_index].height

        for col_idx in range(1, source_ws.max_column + 1):
            source_cell = source_ws.cell(row=row_index, column=col_idx)
            target_cell = target_ws.cell(row=target_row, column=col_idx)

            target_cell.value = source_cell.value
            self._copy_cell_style(source_cell, target_cell)

            col_letter = get_column_letter(col_idx)
            self.column_widths[col_letter] = source_ws.column_dimensions[col_letter].width

    def _copy_images(self, source_ws, row_offset=0):
        """复制图片并保存位置信息 (修复版)"""
        for img in source_ws._images:
            # 保存图片二进制数据
            img_data = img._data

            # 解析图片位置
            if isinstance(img.anchor, str):
                # 处理字符串锚点 (旧格式)
                match = re.match(r'([A-Z]+)(\d+)', img.anchor)
                if match:
                    col = match.group(1)
                    row = int(match.group(2))
                    new_row = row + row_offset
                    anchor_str = f"{col}{new_row}"
                else:
                    anchor_str = img.anchor
                self.images.append({
                    'img_data': img_data,
                    'anchor': anchor_str,
                    'anchor_type': 'string'
                })
            else:
                # 处理对象锚点 (新格式)
                anchor_obj = img.anchor
                if isinstance(anchor_obj, OneCellAnchor):
                    # 获取原始位置
                    from_col = anchor_obj._from.col
                    from_row = anchor_obj._from.row

                    # 保存调整后的位置
                    self.images.append({
                        'img_data': img_data,
                        'anchor': {
                            'from_col': from_col,
                            'from_row': from_row + row_offset,
                            'to_col': from_col,  # 单点锚点
                            'to_row': from_row + row_offset  # 单点锚点
                        },
                        'anchor_type': 'object'
                    })
                else:
                    # 处理其他锚点类型 (如双点锚点)
                    self.images.append({
                        'img_data': img_data,
                        'anchor': img.anchor,
                        'anchor_type': 'original'
                    })

    def _apply_images(self):
        """添加所有图片到合并后的工作表 (修复版)"""
        for img_info in self.images:
            img = Image(BytesIO(img_info['img_data']))

            # 根据锚点类型处理位置
            if img_info['anchor_type'] == 'string':
                img.anchor = img_info['anchor']
            elif img_info['anchor_type'] == 'object':
                # 创建新锚点对象
                anchor_data = img_info['anchor']
                from_marker = AnchorMarker(
                    col=anchor_data['from_col'],
                    row=anchor_data['from_row'],
                    colOff=0,
                    rowOff=0
                )
                to_marker = AnchorMarker(
                    col=anchor_data['to_col'],
                    row=anchor_data['to_row'],
                    colOff=0,
                    rowOff=0
                )
                img.anchor = OneCellAnchor(_from=from_marker, to=to_marker)
            else:
                img.anchor = img_info['anchor']

            self.merged_ws.add_image(img)

    def _get_excel_files(self, folder_path: str) -> List[str]:
        """获取文件夹中所有Excel文件"""
        excel_files = []
        for file in os.listdir(folder_path):
            if file.endswith(('.xls', '.xlsx')):
                if not file.startswith('merged_'):
                    excel_files.append(os.path.join(folder_path, file))
        return sorted(excel_files)

    def _convert_xls_to_xlsx(self, xls_path: str) -> str:
        """将xls文件转换为临时xlsx文件以便处理"""
        try:
            temp_dir = tempfile.mkdtemp()
            file_name = os.path.basename(xls_path)
            xlsx_path = os.path.join(temp_dir, f"{Path(file_name).stem}.xlsx")
            shutil.copy(xls_path, xlsx_path)
            return xlsx_path
        except Exception as e:
            print(f"转换xls文件时出错: {e}")
            raise

    def merge_files(self, folder_path: str):
        """合并文件夹中所有Excel文件"""
        try:
            excel_files = self._get_excel_files(folder_path)
            if not excel_files:
                raise ValueError(f"在文件夹 {folder_path} 中没有找到Excel文件")

            for file_path in excel_files:
                is_temp_file = False
                if file_path.endswith('.xls'):
                    file_path = self._convert_xls_to_xlsx(file_path)
                    is_temp_file = True

                wb = load_workbook(file_path)
                ws = wb.active

                if not self.header_copied:
                    self._copy_row(ws, 1, self.merged_ws, 1)
                    self.header_copied = True
                    self.current_row += 1

                row_offset = self.current_row - 2

                for row_idx in range(2, ws.max_row + 1):
                    self._copy_row(ws, row_idx, self.merged_ws, self.current_row)
                    self.current_row += 1

                self._copy_images(ws, row_offset)

                if is_temp_file:
                    os.remove(file_path)
                    temp_dir = os.path.dirname(file_path)
                    if os.path.exists(temp_dir):
                        shutil.rmtree(temp_dir)

            for col_letter, width in self.column_widths.items():
                if width is not None:
                    self.merged_ws.column_dimensions[col_letter].width = width

            self._apply_images()

        except Exception as e:
            print(f"合并文件时出错: {e}")
            raise

    def save_merged_file(self, output_folder: str):
        """保存合并后的文件"""
        try:
            if not os.path.exists(output_folder):
                os.makedirs(output_folder)

            output_path = os.path.join(output_folder, 'merged_result.xlsx')
            self.merged_wb.save(output_path)
            print(f"合并完成！文件已保存为: {output_path}")
            return output_path
        except Exception as e:
            print(f"保存文件时出错: {e}")
            raise


# 以下main函数保持不变...

def main():
    try:
        merger = ExcelMerger()
        folder_path = r"C:\Users\Administrator\Desktop\上架表格"

        print("开始合并Excel文件...")
        merger.merge_files(folder_path)
        output_path = merger.save_merged_file(folder_path)
        print(f"成功创建合并文件: {output_path}")

    except Exception as e:
        print(f"程序执行出错: {e}")


if __name__ == "__main__":
    main()