import openpyxl
import shutil
import os
import pandas as pd
from datetime import datetime


class AmazonTemplateFiller:
    """
    负责亚马逊上架模板（或其他复杂表头Excel）数据填充的类
    工作流：传入路径 -> 自动备份 -> 初始化加载 -> 循环填充多列 -> 一次性保存
    """

    def __init__(self, original_path, file_index=None, sheet_name='Template', header_row=4, start_row=7):
        """
        初始化：工具会自动复制副本到 [按当天日期动态生成的文件夹]，并支持传入数字生成两位数的后缀编号。

        :param original_path: 原模板文件路径
        :param file_index: 传入数字 (如 1, 2, 10)，内部会自动转为 "01", "02", "10"
        """
        self.original_path = original_path
        self.sheet_name = sheet_name
        self.header_row = header_row
        self.start_row = start_row

        # --- 动态建档修改部分开始 ---
        # 1. 获取当前日期，格式为 YYMMDD (例如 240520)
        date_str = datetime.now().strftime('%y%m%d')

        # 2. 动态定义目标保存目录
        target_dir = os.path.join(r'F:\上架', date_str)

        # 如果该目录不存在，则自动创建
        if not os.path.exists(target_dir):
            os.makedirs(target_dir)
            print(f"📁 自动创建今日工作目录: {target_dir}")

        # 3. 提取原文件名并拼接新文件名
        base_name = os.path.basename(original_path)
        file_name, file_ext = os.path.splitext(base_name)

        # 4. 组合成新的完整路径
        if file_index is not None:
            # 💡 核心细节优化：:02d 的意思是，把它当做整数处理，并且至少保留两位，不足的左边补0
            # 1 -> "01", 9 -> "09", 10 -> "10", 99 -> "99"
            suffix_str = f"{int(file_index):02d}"
            new_filename = f"{file_name}_{date_str}_{suffix_str}{file_ext}"
        else:
            new_filename = f"{file_name}_{date_str}{file_ext}"

        self.new_path = os.path.join(target_dir, new_filename)
        # --- 动态建档修改部分结束 ---

        try:
            # 5. 复制物理文件
            shutil.copy2(self.original_path, self.new_path)
            print(f"✅ 文件已准备就绪")
            print(f"   源文件: {self.original_path}")
            print(f"   工作副本: {self.new_path}")

            # 6. 将副本加载到内存中
            self.wb = openpyxl.load_workbook(self.new_path)

            if self.sheet_name not in self.wb.sheetnames:
                raise ValueError(f"错误：未找到名为 '{self.sheet_name}' 的工作表")

            self.sheet = self.wb[self.sheet_name]

            # 7. 建立列名映射字典
            self.col_map = {}
            for cell in self.sheet[self.header_row]:
                if cell.value:
                    col_name = str(cell.value).strip()
                    self.col_map[col_name] = cell.column

        except Exception as e:
            print(f"❌ 初始化失败: {e}")
            raise e


    def fill_column_data(self, target_col_name, data_list):
        """
        核心方法：填入一整列数据。
        你可以在外部无限次调用这个方法，它都不会报错，也不会马上保存。

        :param target_col_name: 模板里的列名 (例如 "SKU", "item_name")
        :param data_list: 要填入的数据列表 (List)
        """
        # 判断我们要找的列名在不在刚刚扫描的字典里
        if target_col_name not in self.col_map:
            print(f"⚠️ 警告：在第 {self.header_row} 行未找到名为 '{target_col_name}' 的列，本次跳过。")
            return

        # 获取具体的列号 (比如 A列是1, B列是2)
        col_index = self.col_map[target_col_name]
        print(f"⏳ 正在填充列: [{target_col_name}] (共 {len(data_list)} 条数据)...")

        # 遍历你的列表，逐个填入单元格
        for i, value in enumerate(data_list):
            current_row = self.start_row + i
            # 如果值为 None 或者 pandas的 NaN，可以不填
            if pd.isna(value) if 'pd' in globals() and isinstance(value, float) else value is not None:
                self.sheet.cell(row=current_row, column=col_index).value = value

    def save_and_close(self):
        """
        所有列都填完后，最后调用一次这个方法，将内存里的数据真正写进硬盘里。
        """
        try:
            self.wb.save(self.new_path)
            self.wb.close()
            print(f"💾 大功告成！所有修改已保存至: {self.new_path}")
        except PermissionError:
            print("❌ 保存失败：文件可能正被其他程序打开，请先关闭该 Excel。")
        except Exception as e:
            print(f"❌ 保存时发生未知错误: {e}")

    def fill_column_skip_first(self, target_col_name, data_list):
        """
        额外方法：填入一整列数据，但【强制跳过第一个数据】。
        常用于亚马逊上传表：父体（第一行）不需要填价格、数量、尺寸等字段时。

        :param target_col_name: 模板里的列名
        :param data_list: 要填入的数据列表
        """
        # 1. 检查列名是否存在
        if target_col_name not in self.col_map:
            print(f"⚠️ 警告：未找到名为 '{target_col_name}' 的列，跳过填充。")
            return

        col_index = self.col_map[target_col_name]
        print(f"⏳ 正在填充列(跳过首行): [{target_col_name}]...")

        # 2. 核心逻辑：使用 range(1, len(data_list))，直接从索引 1（第二个元素）开始循环
        for i in range(1, len(data_list)):
            value = data_list[i]
            current_row = self.start_row + i  # 因为 i 是从 1 开始的，所以会自动填在第 9 行 (8+1)

            # 过滤掉 None
            if value is not None:
                self.sheet.cell(row=current_row, column=col_index).value = value

    def fill_2d_data_skip_first(self, start_col_name, data_2d_list):
        """
        专门处理多维（二维）列表填充，并【跳过第一行数据】。
        以 start_col_name 为起点，向右依次填充子列表的内容。

        :param start_col_name: 起始列的名称（例如 "Main Image URL" 或 "main_image_url"）
        :param data_2d_list: 二维数据列表，例如 [["url2", "url3"], ["url2", "url3"]]
        """
        # 1. 寻找起始列的索引
        if start_col_name not in self.col_map:
            print(f"⚠️ 警告：未找到起始列 '{start_col_name}'，跳过二维数据填充。")
            return

        start_col_index = self.col_map[start_col_name]
        print(f"⏳ 正在从列 [{start_col_name}] 开始，向右填充二维数据块 (跳过首行)...")

        # 2. 遍历外层列表（代表行），使用 range(1, len) 直接跳过索引为 0 的第一行
        for i in range(1, len(data_2d_list)):
            row_data = data_2d_list[i]  # 拿到当前行的一维列表，例如 ["url2", "url3", "url1"]
            current_row = self.start_row + i

            # 3. 遍历内层列表（代表向右填充的列）
            for j, value in enumerate(row_data):
                # pandas 中空值转为 list 后可能是 nan，这里做一下安全过滤
                # 只有当 value 有效时才写入单元格
                if value is not None and str(value).lower() != 'nan':
                    # 核心定位：行号向下增加，列号向右增加 (start_col_index + j)
                    self.sheet.cell(row=current_row, column=start_col_index + j).value = value

    def fill_2d_data_backward(self, anchor_col_name, data_2d_list):
        """
        以最右侧列为锚点向左填充，且保持 Excel 的视觉顺序与列表定义顺序一致，包含首行。

        :param anchor_col_name: 起始列（最右侧的列）的名称
        :param data_2d_list: 二维数据列表
        """
        if anchor_col_name not in self.col_map:
            print(f"⚠️ 警告：未找到锚点列 '{anchor_col_name}'，跳过逆向填充。")
            return

        anchor_col_index = self.col_map[anchor_col_name]
        print(f"⏳ 正在从列 [{anchor_col_name}] 开始，向左对齐填充数据块 (保持原有视觉顺序)...")

        for i in range(len(data_2d_list)):
            row_data = data_2d_list[i]
            current_row = self.start_row + i

            # 💡 核心修复点：将当前行的数据反转 [::-1]
            # 这样列表末尾的 "71%..." 就会被第一个拿到，填在最右侧的 anchor_col
            # 前面的 "Polyester" 会依次往左侧填。
            # 最终在 Excel 里的视觉顺序就会和你代码里写的顺序一模一样！
            reversed_row_data = row_data[::-1]

            for j, value in enumerate(reversed_row_data):
                target_col = anchor_col_index - j

                # 安全保护
                if target_col < 1:
                    print(f"⚠️ 警告：第 {current_row} 行向左填充时触碰到表格最左侧边缘！")
                    break

                if value is not None and str(value).lower() != 'nan':
                    self.sheet.cell(row=current_row, column=target_col).value = value

# ==========================================
# 测试代码块：演示它有多方便
# ==========================================
# if __name__ == "__main__":
#     # --- 假设我们有以下处理好的数据 (这些通常来自你的其他模块) ---
#     mock_skus = ["SKU-001", "SKU-002", "SKU-003", "SKU-004"]
#     mock_titles = ["This is Title 1", "This is Title 2", "This is Title 3", "This is Title 4"]
#     mock_prices = [19.99, 25.00, 15.50, 9.99]
#
#     # --- 开始执行标准化填充流程 ---
#     template_path = r'C:\Users\Administrator\Desktop\new_yo.xlsx'
#
#     # 1. 实例化 (只写一次路径)
#     filler = AmazonTemplateFiller(original_path=template_path)
#
#     # 2. 疯狂调用填充方法 (只传列名和列表)
#     # 它会自动去第 4 行找这些字眼，然后从第 8 行往下填
#     filler.fill_column_data("SKU", mock_skus)
#     filler.fill_column_data("item_name", mock_titles)  # 假设你的标题列表头叫 item_name
#     filler.fill_column_data("standard_price", mock_prices)
#
#     # 3. 最后统一保存
#     filler.save_and_close()