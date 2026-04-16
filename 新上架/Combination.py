import pandas as pd
import openpyxl
import random
import re
import os


class TitleCombiner:
    """
    专门负责标题拼接、ST关键词填充、大描述生成及重复词清洗的类
    """

    def __init__(self):
        # 默认词表路径配置
        self.path_de = r'C:\Users\Administrator\Desktop\词表DE.xlsx'
        self.path_new = r'C:\Users\Administrator\Desktop\new 词表.xlsx'

        # --- 智能替换词库 (用于标题) ---
        self.replacement_rules = {
            "pants": {3: "pantalones", 4: "pantalones", 5: "trousers", 6: "trousers", 7: "clothes", 8: "clothes",
                      9: "outfits"},
            "shorts": {3: "bermudas", 4: "bermudas", 5: "clothes", 6: "clothes", 7: "outfits"},
            "skirts": {3: "afueras", 4: "afueras", 5: "clothes", 6: "clothes", 7: "outfits"},
            "jeans": {3: "pants", 4: "pants", 5: "trousers", 6: "trousers", 7: "clothes", 8: "clothes", 9: "outfits"},
            "overalls": {3: "overall", 4: "overall", 5: "jumpsuits", 6: "jumpsuits", 7: "outfits", 8: "outfits",
                         9: "rompers", 10: "rompers"},
            "jumpsuits": {3: "jumpsuit", 4: "jumpsuit", 5: "overalls", 6: "overalls", 7: "outfits", 8: "outfits",
                          9: "rompers", 10: "rompers"},
            "rompers": {3: "romper", 4: "romper", 5: "jumpsuits", 6: "jumpsuits", 7: "outfits", 8: "outfits",
                        9: "overalls", 10: "overalls"}
        }

    def smart_replace_and_clean(self, text):
        """核心逻辑：标题去重与替换"""
        text = text.lower()
        words = text.split()
        word_counts = {}
        result = []

        for word in words:
            raw_count = word_counts.get(word, 0) + 1
            word_counts[word] = raw_count
            target_word = word

            if word in self.replacement_rules:
                if raw_count in self.replacement_rules[word]:
                    target_word = self.replacement_rules[word][raw_count]

            final_result_text = " ".join(result)
            final_word_count = final_result_text.split().count(target_word)

            if final_word_count < 2:
                result.append(target_word)
            else:
                continue

        return ' '.join(result)

    def _get_keywords_from_excel(self, flag, cibiao):
        """内部辅助方法：加载 Excel 词表数据"""
        file_path = self.path_de if flag == "LONG5" else self.path_new
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"未找到词表文件: {file_path}")

        wb = openpyxl.load_workbook(file_path, data_only=True)
        if cibiao not in wb.sheetnames:
            raise ValueError(f"词表中不存在名为 [{cibiao}] 的工作表")

        ws = wb[cibiao]
        texts = [str(cell.value).strip() for cell in ws['C'][1:] if cell.value is not None]
        return texts

    # ==========================================
    # 方法 1：生成标题
    # ==========================================
    def create_combined_titles(self, num, cibiao, title_list, colors_series, flag):
        keywords_pool = self._get_keywords_from_excel(flag, cibiao)
        random_strings = []

        # 定义需要小写的介词、连词、冠词
        lowercases = {"with", "and", "of", "in", "to", "for", "on", "at", "by", "a", "an", "the", "or", "from"}

        def apply_custom_case(text):
            """处理大小写：首字母大写，介词小写"""
            if not text:
                return ""
            words = text.split()
            result = []
            for i, word in enumerate(words):
                # 第一个单词，或者不在小写清单里的单词 -> 首字母大写
                # 使用 word[0].upper() + word[1:] 可以保留类似 "USB" 这种全大写缩写
                if i == 0 or word.lower() not in lowercases:
                    if len(word) > 1:
                        result.append(word[0].upper() + word[1:])
                    else:
                        result.append(word.upper())
                else:
                    # 在清单里的介词/连词 -> 转小写
                    result.append(word.lower())
            return " ".join(result)

        for idx in range(num):
            # 基础标题处理
            base_title = str(title_list[idx]).strip() if idx < len(title_list) else ""

            if idx == 0:
                # --- 父标题逻辑：不拼接，只清洗并转换大小写 ---
                cleaned_text = self.smart_replace_and_clean(base_title)
                final_string = apply_custom_case(cleaned_text[:200])
            else:
                # --- 子标题逻辑：拼接关键词和颜色 ---
                color_str = str(colors_series.iloc[idx]).strip() if idx < len(colors_series) else ""

                # 1. 计算剩余空间并抽取关键词
                max_len_for_keywords = 200 - len(base_title) - len(color_str) - 10
                current_keywords = []
                if max_len_for_keywords > 0 and keywords_pool:
                    shuffled_pool = random.sample(keywords_pool, min(len(keywords_pool), 15))  # 限制抽取数提高效率
                    current_len = 0
                    for word in shuffled_pool:
                        if current_len + len(word) + 1 <= max_len_for_keywords:
                            current_keywords.append(word)
                            current_len += len(word) + 1

                keyword_str = " ".join(current_keywords)

                # 2. 拼接
                combined_text = f"{base_title} {keyword_str} {color_str}"

                # 3. 清洗特殊符号
                cleaned_text = self.smart_replace_and_clean(combined_text)

                # 4. 长度截断（确保颜色在最后）
                if len(cleaned_text) > 200:
                    allowed_mid_len = 200 - len(color_str) - 1
                    temp_text = cleaned_text[:allowed_mid_len].rstrip() + " " + color_str
                else:
                    temp_text = cleaned_text

                # 5. 最后统一应用大小写转换
                final_string = apply_custom_case(temp_text)

            random_strings.append(final_string)

        return random_strings

    # ==========================================
    # 方法 2：生成 Search Terms
    # ==========================================
    def create_search_terms(self, num, cibiao, title_list, attr_pool, flag):
        keywords_pool = self._get_keywords_from_excel(flag, cibiao)
        st_strings = []

        for idx in range(num):
            # 基础标题直接转小写
            base_title = str(title_list[idx]).strip().lower()

            # 如果标题本身就超过250，直接截断返回
            if len(base_title) >= 250:
                st_strings.append(base_title[:250])
                continue

            # 随机打乱属性池和关键词池
            shuffled_attrs = random.sample(attr_pool, len(attr_pool)) if attr_pool else []
            shuffled_keywords = random.sample(keywords_pool, len(keywords_pool)) if keywords_pool else []

            current_words = []
            # 当前长度 = 基础标题长度 + 1个空格
            current_len = len(base_title) + 1

            a_idx = 0
            k_idx = 0

            # 交替添加属性词和关键词，直到达到250字符上限
            while a_idx < len(shuffled_attrs) or k_idx < len(shuffled_keywords):
                # 尝试添加属性词
                if a_idx < len(shuffled_attrs):
                    attr_word = str(shuffled_attrs[a_idx]).strip().lower()  # 强制小写
                    if current_len + len(attr_word) <= 250:
                        current_words.append(attr_word)
                        current_len += len(attr_word) + 1
                    a_idx += 1

                # 尝试添加关键词
                if k_idx < len(shuffled_keywords):
                    key_word = str(shuffled_keywords[k_idx]).strip().lower()  # 强制小写
                    if current_len + len(key_word) <= 250:
                        current_words.append(key_word)
                        current_len += len(key_word) + 1
                    k_idx += 1

            # 拼接基础标题和后续补充词
            if current_words:
                combined_text = f"{base_title} {' '.join(current_words)}"
            else:
                combined_text = base_title

            # 最终确保：全小写、去首尾空格、截断至250字符
            final_st = combined_text.lower().strip()[:250]
            st_strings.append(final_st)

        return st_strings

    # ==========================================
    # 方法 3：生成产品描述 (大描述 HTML)
    # ==========================================
    def create_html_description(self, file_path, size_chart, p5):
        """
        根据 HTML 模板、尺码表、五点列表，生成带排版的 HTML 大描述。
        【新增逻辑】：动态检测长度，确保最终生成的字符数严格 <= 2000。

        :param file_path: HTML 模板文件的绝对路径
        :param size_chart: 尺码表字符串 (如果是空值传 pd.NA 或 np.nan)
        :param p5: 包含 5 个五点描述的 List (例如: ["Title: Desc", "Title2: Desc2"])
        :return: 最终拼接完成的 HTML 字符串
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"未找到 HTML 模板文件: {file_path}")

        with open(file_path, 'r', encoding='utf-8') as file:
            html_template = file.read()

        # 1. 净化和解析五点列表 (p5)
        parsed_bullets = []
        for bullet in p5:
            if pd.isna(bullet) or not str(bullet).strip():
                continue
            bp_str = str(bullet).replace("：", ":")
            parts = bp_str.split(":", 1)

            if len(parts) == 2:
                parsed_bullets.append((parts[0].strip(), parts[1].strip()))
            else:
                parsed_bullets.append(("Feature", parts[0].strip()))

        # 2. 判断是否有尺码表
        has_size_chart = pd.notna(size_chart) and str(size_chart).strip() != ""

        # 3. 核心模板替换逻辑
        if has_size_chart:
            formatted_sc = str(size_chart).replace("\n", " <br>\n")
            html_template = re.sub(
                r'(<p><strong>Size Chart:</strong></p>\s*<p>)(.*?)(</p>)',
                rf'\1\n{formatted_sc}\n\3',
                html_template,
                flags=re.IGNORECASE | re.DOTALL
            )
            bullets_to_use = parsed_bullets[:3]
        else:
            html_template = re.sub(
                r'<p><strong>Size Chart:</strong></p>\s*<p>.*?</p>',
                '',
                html_template,
                flags=re.IGNORECASE | re.DOTALL
            )
            bullets_to_use = parsed_bullets[:5]

        # 4. 【新增】：动态拼接五点，进行 2000 字符的长度风控
        MAX_LENGTH = 2000
        base_template_len = len(html_template)
        bullets_html = ""

        for title, desc in bullets_to_use:
            # 构造当前这 1 个五点的 HTML 字符串
            single_bullet_html = f"<p><strong>【{title}】</strong> : {desc}</p>\n"

            # 预计算：已拼接的五点长度 + 当前这个五点的长度 + 模板长度 + 1(预留换行符)
            predicted_length = len(bullets_html) + len(single_bullet_html) + base_template_len + 1

            # 如果拼接后不超过 2000，才允许加进去
            if predicted_length <= MAX_LENGTH:
                bullets_html += single_bullet_html
            else:
                # 一旦发现超限，直接终止循环，放弃后面的五点
                break

        # 5. 组合最终结果
        if bullets_html:
            final_description = bullets_html + "\n" + html_template
        else:
            final_description = html_template

        # 【终极保险】：如果模板本身（或尺码表太大）已经超过 2000 字符，执行强制截断
        if len(final_description) > MAX_LENGTH:
            final_description = final_description[:MAX_LENGTH]

        return final_description


# ==========================================
# 测试代码块
# ==========================================
if __name__ == "__main__":
    combiner = TitleCombiner()

    # 模拟数据
    template_path = "description_template.txt"  # 需要自己建一个这文件测试

    # 模拟创建一下测试 HTML 模板
    with open(template_path, "w", encoding="utf-8") as f:
        f.write("<p><strong>⚠️Important Notice:</strong></p>\n")
        f.write("<p>Before making a purchase...</p>\n")
        f.write("<p><strong>Size Chart:</strong></p>\n<p>\n</p>\n")
        f.write("<p><strong>Note:</strong> recommend choosing larger size.</p>")

    mock_size_chart = "Size: S Waist: 80cm\nSize: M Waist: 84cm"
    # mock_size_chart = pd.NA # 打开这行测试无尺码表的情况

    mock_p5 = [
        "WATERPROOF FABRIC: Protects you from rain.",
        "MULTI POCKETS：Keep your items safe.",
        "STRETCH MOBILITY: Flexible for outdoor.",
        "HIKING GEAR: Best choice.",
        "COLOR VARIETY: Many colors."
    ]

    try:
        html_result = combiner.create_html_description(
            file_path=template_path,
            size_chart=mock_size_chart,
            p5=mock_p5
        )
        print("\n--- HTML 描述生成结果 ---")
        print(html_result)

    except Exception as e:
        print(f"运行失败: {e}")