import os
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import date

# --- 配置区域 ---
folder_path = "F:\\上架\\260414"

# 模式选择：
# 'exclude_merge' -> 排除文件名带 "merge" 的文件
# 'only_merge'    -> 只汇总文件名带 "merge" 的文件
filter_mode = 'exclude_merge'

# 获取今天的日期
today = date.today()
date_str = today.strftime("%Y%m%d")

folder_name = os.path.basename(folder_path)
print(f"正在处理文件夹: {folder_name}")

# 新建 .xlsx 文件路径
xlsx_file_path = f"C:\\Users\\Administrator\\Desktop\\upload_sku\\sku{folder_name}.xlsx"

# 创建新的 Workbook 对象
wb_dest = Workbook()
ws_dest = wb_dest.active
ws_dest.title = 'Merged Data'

# 添加标题行
ws_dest["A1"] = "系统SKU"
ws_dest["B1"] = "平台SKU"

original_data = []
processed_data = []

# --- 遍历文件夹 ---
for filename in os.listdir(folder_path):
    if filename.startswith('~$'):
        continue

    if not filename.endswith('.xlsx'):
        continue

    is_merge_file = "merge" in filename.lower()

    if filter_mode == 'exclude_merge' and is_merge_file:
        print(f"跳过文件 (排除模式): {filename}")
        continue
    elif filter_mode == 'only_merge' and not is_merge_file:
        print(f"跳过文件 (仅匹配模式): {filename}")
        continue

    file_path = os.path.join(folder_path, filename)
    print(f"正在读取: {filename}")

    try:
        wb_src = load_workbook(filename=file_path, read_only=True, data_only=True)

        # 找工作表
        ws_name = None
        if 'Template' in wb_src.sheetnames:
            ws_name = 'Template'
        elif 'Vorlage' in wb_src.sheetnames:
            ws_name = 'Vorlage'

        if ws_name is None:
            print(f"警告: 文件 {filename} 中既找不到 'Template' 也找不到 'Vorlage'")
            wb_src.close()
            continue

        ws_src = wb_src[ws_name]
        print(f"成功找到工作表: [{ws_name}] 在文件: {filename}")

        # === 新逻辑：识别SKU列（第4行） ===
        sku_col_index = None
        header_row = list(ws_src.iter_rows(min_row=4, max_row=4, values_only=True))[0]

        for idx, cell in enumerate(header_row):
            if cell and str(cell).strip().lower() == 'sku':
                sku_col_index = idx + 1  # openpyxl列从1开始
                break

        if sku_col_index is None:
            print(f"警告: 文件 {filename} 未找到 SKU 列")
            wb_src.close()
            continue

        print(f"识别到 SKU 列: 第 {sku_col_index} 列")

        # === 从第7行开始读取数据 ===
        file_count = 0

        for row in ws_src.iter_rows(min_row=7, values_only=True):
            if row and len(row) >= sku_col_index and row[sku_col_index - 1] is not None:
                sku = str(row[sku_col_index - 1]).strip()
                original_data.append(sku)

                # SKU 清洗逻辑
                if sku.startswith(('V', 'T', 'C', 'Z', 'P')):
                    sku = sku[20:]
                elif sku.startswith(('li', 'yo', 'lc')):
                    sku = sku[10:]

                processed_data.append(sku)
                file_count += 1

        print(f"本文件识别SKU数量: {file_count}")
        print(f"示例SKU(前三个): {original_data[:3]}")

        wb_src.close()

    except Exception as e:
        print(f"处理文件 {filename} 时出错: {e}")

# --- 写入数据 ---
for i, sku in enumerate(processed_data):
    ws_dest.cell(row=i + 2, column=1, value=sku)

for i, sku in enumerate(original_data):
    ws_dest.cell(row=i + 2, column=2, value=sku)

# 保存文件
try:
    os.makedirs(os.path.dirname(xlsx_file_path), exist_ok=True)
    wb_dest.save(filename=xlsx_file_path)
    print(f"处理完成！文件已保存至: {xlsx_file_path}")

    # === 总统计 ===
    print(f"\n总SKU数量: {len(original_data)}")
    print(f"总示例SKU(前三个): {original_data[:3]}")

except Exception as e:
    print(f"保存文件失败: {e}")

wb_dest.close()