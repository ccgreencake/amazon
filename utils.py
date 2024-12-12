import math
import pandas as pd
from openpyxl import load_workbook
import os
import subprocess
import numpy as np
def description(file_path,size_chart,p5):
    # 打开文本文件进行读取
    with open(file_path, 'r', encoding='utf-8') as file:
        # 读取文件内容
        file_content = file.read()
    i = 0
    while True:
        if type(p5[i]) == float and math.isnan(p5[i]):
            break
        i += 1
    part = []
    for n in range(i, 0, -1):
        part.append(p5[n - 1].split(":"))
    product_description = ''
    position_strong = {}
    position = {}
    for n in range(i):
        if n == 0:
            position_strong[n] = file_content.find("<p><strong></strong></p>")
            position[n] = file_content.find("<p></p>")
        else:
            position_strong[n] = file_content.find("<p><strong></strong></p>", position_strong[n - 1] + 1)
            position[n] = file_content.find("<p></p>", position[n - 1] + 1)
    if size_chart!=np.nan and type(size_chart)!= np.float64:
        output_string = size_chart.replace("\n", " <br>\n")
        # 进行替换操作
        file_content = file_content.replace("<p><strong>Size Chart:</strong></p>\n<p>",
                                            "<p><strong>Size Chart:</strong></p>\n<p>\n" + output_string)
        # first_position = file_content.find("<p><strong></strong></p>")
        # # 查找第二个 <p><strong> 的位置
        # second_position = file_content.find("<p><strong></strong></p>", first_position + 1)
        # # 查找第3个 <p><strong> 的位置
        # third_position = file_content.find("<p><strong></strong></p>", second_position + 1)
        # # 查找第一个 <p></p> 的位置
        # first_p = file_content.find("<p></p>")
        # # 查找第二个 <p></p>  的位置
        # second_p = file_content.find("<p></p>", first_p + 1)
        # # 查找第二个 <p></p>  的位置
        # third_p = file_content.find("<p></p>", second_p + 1)

        for n in range(3, 0, -1):
            product_description = '<p><strong>' + part[n][0] + '</strong></p>'+'<p>' + part[n][1] + '</p>' + product_description

        product_description = product_description + file_content
        return product_description
    elif np.isnan(size_chart):
        # 进行替换操作
        file_content = file_content.replace("<p><strong>Size Chart:</strong></p>\n<p>\n</p>",
                                            "")


        # product_description = file_content[:third_p] + "<p>" + part3[1] + "</p>" + file_content[third_p + 7:]
        # product_description = product_description[:third_position] + "<p><strong>" + part3[
        #     0] + "</strong></p>" + product_description[third_position + 24:]
        # product_description = product_description[:second_p] + "<p>" + part2[1] + "</p>" + product_description[
        #                                                                                    second_p + 7:]
        # product_description = product_description[:second_position] + "<p><strong>" + part2[
        #     0] + "</strong></p>" + product_description[second_position + 24:]
        # product_description = product_description[:first_p] + "<p>" + part1[1] + "</p>" + product_description[
        #                                                                                   first_p + 7:]
        # product_description = product_description[:first_position] + "<p><strong>" + part1[
        #     0] + "</strong></p>" + product_description[first_position + 24:]
        for n in range(5, 0, -1):
            product_description = '<p><strong>' + part[n][0] + '</strong></p>' + '<p>' + part[n][
                1] + '</p>' + product_description

        product_description = product_description + '\n'+file_content
        return product_description
    # else:
    #     output_string = size_chart.replace("\n", " <br>\n")
    #     # 打开文本文件进行读取
    #     with open(file_path, 'r', encoding='utf-8') as file:
    #         # 读取文件内容
    #         file_content = file.read()
    #     # 进行替换操作
    #     file_content = file_content.replace("<p><strong>Size Chart:</strong></p>\n<p>", "<p><strong>Size Chart:</strong></p>\n<p>\n" + output_string)
    #     first_position = file_content.find("<p><strong></strong></p>")
    #     # 查找第二个 <p><strong> 的位置
    #     second_position = file_content.find("<p><strong></strong></p>", first_position + 1)
    #     # 查找第3个 <p><strong> 的位置
    #     third_position = file_content.find("<p><strong></strong></p>", second_position + 1)
    #     # 查找第一个 <p></p> 的位置
    #     first_p = file_content.find("<p></p>")
    #     # 查找第二个 <p></p>  的位置
    #     second_p = file_content.find("<p></p>", first_p + 1)
    #     # 查找第二个 <p></p>  的位置
    #     third_p = file_content.find("<p></p>", second_p + 1)
    #
    #     product_description = file_content[:third_p] + "<p>" + part3[1] + "</p>" + file_content[third_p + 7:]
    #     product_description = product_description[:third_position] + "<p><strong>" + part3[0] + "</strong></p>" + product_description[third_position + 24:]
    #     product_description = product_description[:second_p] + "<p>" + part2[1] + "</p>" + product_description[second_p + 7:]
    #     product_description = product_description[:second_position] + "<p><strong>" + part2[0] + "</strong></p>" + product_description[second_position + 24:]
    #     product_description = product_description[:first_p] + "<p>" + part1[1] + "</p>" + product_description[first_p + 7:]
    #     product_description = product_description[:first_position] + "<p><strong>" + part1[0] + "</strong></p>" + product_description[first_position + 24:]
    #     return product_description

def rise(string):
    if "High" in string:
        rise = 'high'
    elif "Low" in string:
        rise = 'low'
    else:
        rise = 'mid'
    return rise

def item_type(string,feed_product):
    if feed_product == 'pants':
        if "Dress" in string:
            item_type = 'dress-pants'
        elif "Yoga" in string:
            item_type = 'yoga-pants'
        elif "Casual" in string:
            item_type = 'casual-pants'
        else:
            item_type = 'pants'
    elif feed_product == 'shorts':
        if "denim" in string or "jeans" in string:
            item_type = 'denim-shorts'
        elif "cargo" in string:
            item_type = 'cargo-shorts'
        elif "running" in string or "jogger" in string:
            item_type = 'running-shorts'
        elif "hiking" in string:
            item_type = 'hiking-shorts'
        else:
            item_type = 'shorts'
    elif feed_product == 'onepieceoutfit':
        item_type = 'jumpsuits-apparel'
    elif feed_product == 'overalls':
        item_type = 'overalls'
    else:
        item_type = 'unknown'
    return item_type

def feed_product(string):
    if 'Shorts' in string:
        feed_product = 'shorts'
    elif 'Overalls' in string:
        feed_product = 'overalls'
    elif 'Jumpsuit' in string:
        feed_product = 'onepieceoutfit'
    else:
        feed_product = 'pants'
    return feed_product

def leg_style(string):
    if "Flare" in string:
        leg_style = 'Flared'
    elif "Cropped" in string:
        leg_style = 'Cropped'
    else:
        leg_style = 'Wide'
    return leg_style

def copy_file(file_b_path, file_a_path):
    # 使用pandas读取文件A的Sheet1
    df_a = pd.read_excel(file_a_path, engine='openpyxl')

    # 使用openpyxl加载文件B
    wb = load_workbook(file_b_path)
    ws = wb['Template'] if 'Template' in wb.sheetnames else wb.create_sheet('Template')

    # 设置从第四行开始的所有行的行高为20磅
    for row in range(4, ws.max_row + 1):
        ws.row_dimensions[row].height = 20

    # 读取数据到列表
    data = df_a.values.tolist()

    # 从第四行开始写入数据
    for row_num, row_data in enumerate(data, start=1):  # 从第一行开始写入，因为已经从第四行设置了行高
        for col_num, value in enumerate(row_data, start=1):
            ws.cell(row=row_num + 3, column=col_num, value=value)  # 加3是因为我们从第四行开始设置行高

    # 保存工作簿
    wb.save(file_b_path)
    try:
        os.remove(file_a_path)
        print(f"文件 '{file_a_path}' 已被删除。")
    except OSError as e:
        print(f"删除文件时出错: {e}")
    # 检查操作系统并使用WPS打开文件B
    if os.name == 'nt':  # Windows系统
        # 假设WPS的路径是 "C:\Program Files\WPS Office\et.exe"
        wps_path = r'C:\Users\123\AppData\Local\Kingsoft\WPS Office\ksolaunch.exe'
        # 使用WPS打开Excel文件
        subprocess.call([wps_path, file_b_path])

    print(f"文件 '{file_b_path}' 已使用WPS打开。")