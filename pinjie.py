import openpyxl
import random
import subprocess
import re
import re

excel_file = r'C:\Users\Administrator\Desktop\new 词表.xlsx'
import re

import re


def truncate_string_to_200(input_str):
    # 将字符串按空格分割成单词列表
    words = input_str.split()

    # 如果原始字符串长度小于等于 200，直接返回
    if len(input_str) <= 200:
        return input_str

    # 从末尾开始逐步删除单词，直到长度小于等于 200
    while len(' '.join(words)) > 200:
        words.pop()

    # 将处理后的单词列表重新拼接成字符串
    truncated_str = ' '.join(words)

    return truncated_str

def process_string(input_string):
    # 定义替换规则
    replacement_rules = {
        "pants": {
            5: "trousers",
            6: "trousers",
            3: "pantalones",
            4: "pantalones",
            7: "clothes",
            8: "clothes",
            9: "outfits"
        },
        "shorts": {
            3: "bermudas",
            4: "bermudas",
            5: "clothes",
            6: "clothes",
            7: "outfits"
        },
        "skirts":{
            3: "afueras",
            4: "afueras",
            5: "clothes",
            6: "clothes",
            7: "outfits"
        },
        "jeans": {
            5: "trousers",
            6: "trousers",
            3: "pants",
            4: "pants",
            7: "clothes",
            8: "clothes",
            9: "outfits"
        },
        "overalls":{
            3: "overall",
            4: "overall",
            5: "jumpsuits",
            6: "jumpsuits",
            7: "outfits",
            8: "outfits",
            9: "rompers",
            10: "rompers",

        },
        "jumpsuits": {
            5: "overalls",
            6: "overalls",
            3: "jumpsuit",
            4: "jumpsuit",
            7: "outfits",
            8: "outfits",
            9: "rompers",
            10: "rompers",

        },
        "rompers": {
            9: "overalls",
            10: "overalls",
            5: "jumpsuits",
            6: "jumpsuits",
            7: "outfits",
            8: "outfits",
            3: "romper",
            4: "romper",

        }
    }

    # 定义需要保留的单词
    reserved_words = {}

    # 将字符串拆分为单词列表
    words = input_string.split()

    # 统计每个单词的出现次数
    word_count = {}

    # 第一次修改：根据规则替换单词
    modified_words = []
    for word in words:
        # 统计单词出现次数
        word_count[word] = word_count.get(word, 0) + 1

        # 如果单词是需要替换的规则单词
        if word in replacement_rules:
            # 获取替换规则
            rule = replacement_rules[word]
            # 获取当前出现次数
            count = word_count[word]
            # 如果当前出现次数在规则中，进行替换
            if count in rule:
                word = rule[count]

        modified_words.append(word)

    # 将修改后的单词列表重新组合成字符串
    modified_string = " ".join(modified_words)

    # 第二次修改：删除出现超过两次的单词，但保留指定单词
    word_count = {}
    final_words = []
    for word in modified_string.split():
        word_count[word] = word_count.get(word, 0) + 1
        if word_count[word] <= 2 or word in reserved_words:
            final_words.append(word)

    # 第三次修改：替换两个空格为一个空格
    final_string = re.sub(r'\s+', ' ', " ".join(final_words))

    return final_string


# # 测试字符串
# input_string = "cargo pants cargo pants cargo tactial pants asdoijo pants oaisdjo pants ajoaisjd pants sweatpants"
# # 调用函数
# result = process_string(input_string)
# print(result)
# 读取Excel文件
wb = openpyxl.load_workbook(excel_file)
shuxing = [
                    "baggy",
                    "casual",
                    "plus size",
                    "comfy",
                    "soft",
                    "comfortable",
                    "elegant",
                    "cozy",
                    "trendy",
                    "loose fit",
                    "summer",
                    "lightweight"
                    "oversized",
                    "breathable",
                    "petite",
                    "y2k"
                ]



def remove_third_and_later_duplicates(text):
    text = text.lower()
    words = text.split()  # 将文本分割为单词列表
    seen = {}  # 用于记录每个单词出现的次数
    result = []  # 用于存储最终结果

    for word in words:
        if word in seen:
            seen[word] += 1
        else:
            seen[word] = 1

        # 如果单词出现次数小于等于2，则保留
        if seen[word] <= 2:
            result.append(word)

    return ' '.join(result)  # 将结果列表重新组合为字符串
# 要处理的Excel文件路径
def zibiao(num, cibiao, title, colors):
    ws = wb[cibiao]
    texts = [cell.value.strip() for cell in ws['C'] if cell.value and isinstance(cell.value, str)]
    texts.pop(0)
    if len(texts) < 1:
        print("工作表中没有足够的文本数据。")
    else:
        random_strings = []
        tiaoguo = 0
        for idx in range(num):
            # 第一个标题单独处理，不拼接正文和颜色
            if idx == 0:
                random_string = title[idx]
                # 也不拼接颜色
                random_strings.append(random_string)
                tiaoguo = 1
                continue

            color_str = colors.iloc[idx]
            color_str = str(color_str).strip()  # 颜色字符串
            max_len_for_text = 200 - len(color_str) - 1  # 留1个空格给颜色前的空格

            random_string = title[idx] + " "
            total_length = len(random_string)
            flag = True

            # 用动态调整后的 max_len_for_text 替代固定200，保证拼文字不突破限制
            while total_length < max_len_for_text and flag:
                for text in random.sample(texts, len(texts)):
                    if total_length + len(text) + 1 <= max_len_for_text:
                        random_string += text + " "
                        total_length += len(text) + 1
                    else:
                        flag = False
                        break

            random_string = re.sub(r'\b(\w+)\s+\1\b', r'\1', random_string)
            random_string = process_string(random_string)

            flag = True
            while total_length < max_len_for_text and flag:
                for text in random.sample(texts, len(texts)):
                    if total_length + len(text) + 1 <= max_len_for_text:
                        random_string += " " + text
                        total_length += len(text) + 1
                    else:
                        flag = False
                        break

            random_string = re.sub(r'\b(\w+)\s+\1\b', r'\1', random_string)

            random_string = remove_third_and_later_duplicates(random_string)
            random_string = truncate_string_to_200(process_string(random_string))  # 保险保证

            # 最后拼接颜色字符串，确保总长度不超200（理论不会超）
            final_string = random_string.rstrip() + " " + color_str
            if len(final_string) > 200:
                # 保险策略，如果超了，再截断正文部分
                allowed_length = 200 - len(color_str) - 1
                final_string = random_string[:allowed_length].rstrip() + " " + color_str

            random_strings.append(final_string)

        return random_strings
def pinjie1(num,cibiao,title,fangfa):
    # excel_file = r'C:\Users\Administrator\Desktop\new 词表.xlsx'
    #
    # # 读取Excel文件
    # wb = openpyxl.load_workbook(excel_file)

    ws = wb[cibiao]
    changdu = 250
    # changdu=500
    # 读取第一列的所有文本
    texts = [cell.value.strip() for cell in ws['A'] if cell.value and isinstance(cell.value, str)]
    texts.pop(0)
    # 检查texts列表是否有足够的文本片段
    if len(texts) < 1:
        print("工作表中没有足够的文本数据。")
    else:
        # 生成num个长度接近250字符的随机字符串
        random_strings = []
        for _ in range(num):
            if fangfa == 0:
                random_string = title[_]+" "
                total_length = len(random_string)
            else :
                random_string = ""
                total_length = 0
            flag = True
            while total_length < changdu and flag :  # 当总长度小于250且还有文本片段时继续
                shuxing = [
                    "baggy",
                    "casual",
                    "plus size",
                    "comfy",
                    "soft",
                    "comfortable",
                    "elegant",
                    "cozy",
                    "trendy",
                    "loose fit",
                    "summer",
                    "lightweight"
                    "oversized",
                    "breathable",
                    "petite",
                    "y2k"
                ]
                random.shuffle(shuxing)
                shuxing = shuxing+shuxing[:]+shuxing[:]+shuxing[:]
                n=0
                for text in random.sample(texts, len(texts)):  # 随机打乱texts顺序
                    if total_length + len(text) + 3 + len(shuxing[n]) + len(shuxing[n+1]) <= changdu:  # +3是为了加上空格
                        random_string += shuxing[n] + " " + shuxing[n+1] + " " + text + " "  # 添加文本片段和空格
                        total_length += len(text) + 3 + len(shuxing[n]) + len(shuxing[n+1])
                        n+=2
                    else:
                        flag= False
                        break  # 退出内层循环，保持当前random_string

            # 去除最后一个多余的空格
            random_string = random_string.strip()
            random_strings.append(random_string)
    return random_strings
        # # 创建一个新的工作簿来保存结果
        # new_wb = openpyxl.Workbook()
        # new_ws = new_wb.active
        #
        # # 将生成的字符串及其长度写入新的工作表
        # for i, random_string in enumerate(random_strings, 1):
        #     new_ws.cell(row=i, column=1, value=random_string)  # 写入字符串
        #     new_ws.cell(row=i, column=2, value=len(random_string))  # 写入字符串长度
        #
        # # 定义新Excel文件的保存路径
        # new_excel_file = r'C:\Users\Administrator\Desktop\random_strings.xlsx'
        # # 保存新的Excel文件
        # new_wb.save(new_excel_file)
        #
        # print(f"文件已保存: {new_excel_file}")
        #
        # # 使用WPS打开Excel文件（如果需要）
        # wps_path = r'C:\Users\Administrator\AppData\Local\Kingsoft\WPS Office\ksolaunch.exe'
        # subprocess.call([wps_path, new_excel_file])
        # print(f"文件 '{new_excel_file}' 已使用WPS打开。")
def pinjie2(num, cibiao, flag):
    ws = wb[cibiao]
    changdu = 500
    texts = [cell.value.strip() for cell in ws['A'] if cell.value and isinstance(cell.value, str)]
    texts.pop(0)
    if len(texts) < 1:
        print("工作表中没有足够的文本数据。")
    else:
        random_strings = []
        for _ in range(num):
            if flag == 'li':
                random_string = "Bakgeerle "
                total_length = 10
            elif flag == 'lc':
                random_string = "lcyhony "
                total_length = 8
            else:
                random_string = ""
                total_length = 0
            flag = True
            while total_length < changdu and flag:
                for text in random.sample(texts, len(texts)):
                    if total_length + len(text) + 1 <= changdu:
                        random_string += text + " "
                        total_length += len(text) + 1
                    else:
                        flag = False
                        break
            random_string = random_string.strip()
            # 第一次处理重复单词
            random_string = re.sub(r'\b(\w+)\s+\1\b', r'\1', random_string)
            flag = True
            # 检查长度是否小于500，如果是，则继续拼接
            if len(random_string) < changdu:
                while total_length < changdu and flag:
                    for text in random.sample(texts, len(texts)):
                        if len(random_string) + len(text) + 1 <= changdu:
                            random_string += " " + text
                            total_length += len(text) + 1
                        else:
                            flag = False
                            break
            random_string = random_string.strip()
            # 第二次处理重复单词
            random_string = re.sub(r'\b(\w+)\s+\1\b', r'\1', random_string)
            random_strings.append(random_string)
    return random_strings

def pinjiedaochu(num,cibiao,flag):
    ws = wb[cibiao]
    # changdu=250
    changdu = 500
    # 读取第一列的所有文本
    texts = [cell.value.strip() for cell in ws['A'] if cell.value and isinstance(cell.value, str)]
    texts.pop(0)
    # 检查texts列表是否有足够的文本片段
    if len(texts) < 1:
        print("工作表中没有足够的文本数据。")
    else:
        # 生成num个长度接近250字符的随机字符串
        random_strings = []
        for _ in range(num):
            if flag == 'li':
                random_string = "Bakgeerle "
                total_length = 10
            elif flag == 'lc':
                random_string = "lcyhony "
                total_length = 8
            else:
                random_string = ""
                total_length = 0
            flag = True
            while total_length < changdu and flag:  # 当总长度小于250且还有文本片段时继续
                for text in random.sample(texts, len(texts)):  # 随机打乱texts顺序
                    if total_length + len(text) + 1 <= changdu:  # +1是为了加上空格
                        random_string += text + " "  # 添加文本片段和空格
                        total_length += len(text) + 1
                    else:
                        flag = False
                        break  # 退出内层循环，保持当前random_string

            # 去除最后一个多余的空格
            random_string = random_string.strip()
            random_strings.append(random_string)
            # 创建一个新的工作簿来保存结果
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active

    # 将生成的字符串及其长度写入新的工作表
    for i, random_string in enumerate(random_strings, 1):
        new_ws.cell(row=i, column=1, value=random_string)  # 写入字符串
        new_ws.cell(row=i, column=2, value=len(random_string))  # 写入字符串长度

    # 定义新Excel文件的保存路径
    new_excel_file = r'C:\Users\Administrator\Desktop\random_strings.xlsx'
    # 保存新的Excel文件
    new_wb.save(new_excel_file)

    print(f"文件已保存: {new_excel_file}")

    # 使用WPS打开Excel文件（如果需要）
    wps_path = r'C:\Users\Administrator\AppData\Local\Kingsoft\WPS Office\ksolaunch.exe'
    subprocess.call([wps_path, new_excel_file])
    print(f"文件 '{new_excel_file}' 已使用WPS打开。")
# pinjiedaochu(10,'weiku','li')