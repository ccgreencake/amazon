import openpyxl
import os
import re


def split_excel_by_column():
    # 配置文件路径和参数
    source_file = r"C:\Users\Administrator\Desktop\wyt\08.xlsx"
    output_dir = r"C:\Users\Administrator\Desktop\wyt\08"
    sheet_name = "Sheet1"

    # 如果输出文件夹不存在，则自动创建
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"已创建文件夹: {output_dir}")

    print("正在加载源表格，请稍候...")
    # 加载工作簿
    wb = openpyxl.load_workbook(source_file, data_only=True)

    # 检查是否存在目标Sheet
    if sheet_name not in wb.sheetnames:
        print(f"错误：找不到名为 '{sheet_name}' 的工作表！")
        return

    ws = wb[sheet_name]

    # 1. 提取前5行作为公共标题行
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        headers.append(row)

    # 2. 从第6行开始读取数据，并按G列（第7列，索引为6）进行分组
    data_groups = {}

    for row in ws.iter_rows(min_row=6, values_only=True):
        # 如果整行都是空的，则跳过
        if all(cell is None for cell in row):
            continue

        # 获取G列的值 (A=0, B=1, ..., G=6)
        g_value = row[6]

        # 处理G列为空的情况
        if g_value is None or str(g_value).strip() == "":
            g_value = "空值"
        else:
            g_value = str(g_value).strip()

        # 将当前行加入对应的分组字典中
        if g_value not in data_groups:
            data_groups[g_value] = []
        data_groups[g_value].append(row)

    print(f"数据读取完毕，共分为 {len(data_groups)} 个不同的表格。正在生成文件...")

    # 3. 为每个分组创建新的表格并保存
    for g_val, rows in data_groups.items():
        # 清理文件名中的非法字符（Windows文件名不能包含 \ / : * ? " < > |）
        safe_filename = re.sub(r'[\\/*?:"<>|]', "_", g_val)

        # 创建新的工作簿
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        new_ws.title = sheet_name

        # 写入前5行标题
        for header_row in headers:
            new_ws.append(header_row)

        # 写入该分组对应的数据行
        for data_row in rows:
            new_ws.append(data_row)

        # 保存文件
        save_path = os.path.join(output_dir, f"{safe_filename}.xlsx")
        new_wb.save(save_path)
        print(f"成功生成: {safe_filename}.xlsx")

    print("全部拆分完成！")


if __name__ == "__main__":
    split_excel_by_column()