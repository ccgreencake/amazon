import openpyxl


def parse_cell_content(cell_content):
    """
    解析单元格内容并提取所需信息
    """
    leixing = ""
    name = "未知"  # 默认值为 "未知"
    money = ""

    # 尝试从内容中提取类型（收入/支出）
    try:
        leixing = cell_content.split("于")[1][8:10]  # 提取“于”后两个字符
    except IndexError:
        leixing = "未知"

    # 尝试提取对方名称
    try:
        name_start = cell_content.index("对方为") + len("对方为")
        name_end = cell_content.index("(账号")
        name = cell_content[name_start:name_end]

        # 检查从第一个“(”开始的后续四个字符

    except ValueError:
        fee_start = cell_content.index("(") + 1
        fee_type = cell_content[fee_start:fee_start + 4]
        if fee_type == "网银收费":
            name = "手续费"
        elif fee_type == "其他批量":
            name = "回单费用" # 如果没有“对方为”或“(账号”，设置为未知
        else:
            name = "未知"

    # 尝试提取金额
    try:
        money_start = cell_content.index("人民币") + len("人民币")
        money_end = cell_content.index("当前余额")
        money = cell_content[money_start:money_end].strip()
        money = float(money.replace(',', ''))  # 去掉逗号并转换为浮点数
    except ValueError:
        money = 0.00  # 如果没有找到金额，默认设置为0

    return leixing, name, money


def process_xlsx(input_file, output_file):
    """
    读取输入的 xlsx 文件，整理数据并写入新的 xlsx 文件
    """
    # 打开输入文件
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active

    # 创建一个新的工作簿
    new_wb = openpyxl.Workbook()
    new_sheet1 = new_wb.active
    new_sheet1.title = "整理后的数据"
    new_sheet2 = new_wb.create_sheet(title="分类汇总")

    # 写入表头
    new_sheet1.append(["Name", "Leixing", "Money", "支取汇总", "收入汇总"])
    new_sheet2.append(["Name", "Leixing", "Total Money","支取汇总", "收入汇总"])

    # 初始化支取和收入的总和
    zhi_qu_total = 0.00
    shou_ru_total = 0.00

    # 初始化分类汇总的字典
    summary_dict = {}

    # 遍历第一列的数据
    for row in sheet.iter_rows(min_col=1, max_col=1):  # 只读取第一列
        cell = row[0]
        if cell.value:  # 确保单元格有值
            leixing, name, money = parse_cell_content(cell.value)
            new_sheet1.append([name, leixing, money, "", ""])
            if leixing == "支取":
                zhi_qu_total += money
            elif leixing == "收入":
                shou_ru_total += money

            # 更新分类汇总的字典
            key = (name, leixing)
            if key in summary_dict:
                summary_dict[key] += money
            else:
                summary_dict[key] = money

    # 写入支取和收入的总和
    new_sheet1.cell(row=2, column=4, value=zhi_qu_total)
    new_sheet1.cell(row=2, column=5, value=shou_ru_total)

    # 写入分类汇总的数据
    for (name, leixing), total_money in summary_dict.items():
        new_sheet2.append([name, leixing, total_money])


    # 按leixing列降序排列
    data = list(new_sheet2.iter_rows(min_row=2, max_col=3, values_only=True))
    new_sheet2.delete_rows(2, new_sheet2.max_row - 1)
    for row in data:
        new_sheet2.append(row)

    # 写入收款合计
    income_data = [row for row in data if row[1] == "收入"]
    output_data = [row for row in data if row[1] == "支取"]
    total_income = sum(row[2] for row in income_data)
    total_output = sum(row[2] for row in output_data)
    income_details = ", ".join(f"{row[0]} {row[2]:.2f}元" for row in income_data)
    output_details = ", ".join(f"{row[0]} {row[2]:.2f}元" for row in output_data)
    summary_message = f"收款合计：{total_income:.2f}元（{income_details}）"
    summary_message2 = f"付款合计：{total_output:.2f}元（{output_details}）"
    new_sheet2.cell(row=1, column=6, value=summary_message)
    new_sheet2.cell(row=1, column=7, value=summary_message2)
    new_sheet2.cell(row=2, column=4, value=zhi_qu_total)
    new_sheet2.cell(row=2, column=5, value=shou_ru_total)

    # 保存到新的文件
    new_wb.save(output_file)
    print(f"整理后的数据已保存到 {output_file}")

# 执行脚本
input_file = r"C:\Users\123\Desktop\工作簿1.xlsx"  # 输入文件名
output_file = r"C:\Users\123\Desktop\nana.xlsx"  # 输出文件名
process_xlsx(input_file, output_file)