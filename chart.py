import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
import re


def process_and_chart_with_filter(filename):
    print(f"开始处理文件: {filename}")
    wb = openpyxl.load_workbook(filename)

    # 检查必要的Sheet是否存在
    if 'Sheet1' not in wb.sheetnames or 'Sheet2' not in wb.sheetnames:
        print("错误：文件必须包含 Sheet1 和 Sheet2")
        return

    sheet1 = wb['Sheet1']
    sheet2 = wb['Sheet2']

    # 1. 获取需要筛选的关键词列表 (Sheet2 A列)
    target_keywords = []
    for row in sheet2.iter_rows(min_row=1, max_col=1, values_only=True):
        if row[0]:
            # 转为字符串并去除前后空格
            target_keywords.append(str(row[0]).strip())

    # 去重，防止重复计算
    target_keywords = list(set(target_keywords))
    print(f"检测到 {len(target_keywords)} 个目标关键词: {target_keywords}")

    # 2. 初始化数据容器
    # 结构: {'tactical': [100, 200, ...], 'keyword2': [50, 60, ...]}
    # 假设数据是第2列到第25列 (共24个月)
    data_col_start = 2
    data_col_end = 25
    num_months = data_col_end - data_col_start + 1

    aggregated_data = {k: [0] * num_months for k in target_keywords}

    # 获取表头 (时间标签)
    header_labels = []
    for col in range(data_col_start, data_col_end + 1):
        val = sheet1.cell(row=1, column=col).value
        header_labels.append(val)

    # 3. 遍历 Sheet1 进行匹配和求和
    # 为了性能，使用 values_only=True
    rows = list(sheet1.iter_rows(min_row=2, values_only=True))

    for row_idx, row_values in enumerate(rows):
        phrase = str(row_values[0]) if row_values[0] else ""  # 第一列是词

        # 提取当前行的数据部分 (注意索引，row_values索引从0开始，excel列从1开始)
        # Excel第2列对应 row_values[1]
        row_data = row_values[1:25]

        # 处理数据中的 None，转为 0
        clean_row_data = [v if (v is not None and isinstance(v, (int, float))) else 0 for v in row_data]

        # 如果数据列长度不足24，补0 (防止报错)
        if len(clean_row_data) < num_months:
            clean_row_data.extend([0] * (num_months - len(clean_row_data)))

        # 针对每个目标词进行正则匹配
        for target in target_keywords:
            # 正则逻辑：\b 表示单词边界，确保匹配 tactical 但不匹配 tacticals
            # re.IGNORECASE 忽略大小写
            pattern = r'\b' + re.escape(target) + r'\b'

            if re.search(pattern, phrase, re.IGNORECASE):
                # 如果匹配成功，将数据累加到对应的关键词总和中
                current_sums = aggregated_data[target]
                aggregated_data[target] = [sum(x) for x in zip(current_sums, clean_row_data)]

    # 4. 将汇总结果写入新的临时工作表 'ProcessedData' 以便画图
    if 'ProcessedData' in wb.sheetnames:
        del wb['ProcessedData']
    ws_data = wb.create_sheet('ProcessedData')

    # 写入表头
    ws_data.cell(row=1, column=1, value="Keyword")
    for idx, label in enumerate(header_labels):
        ws_data.cell(row=1, column=idx + 2, value=label)

    # 写入汇总数据
    current_write_row = 2
    valid_keywords_order = []  # 记录写入的顺序，画图用

    for keyword in target_keywords:
        sums = aggregated_data[keyword]
        # 如果这个词全是0（没有匹配到任何数据），可以选择跳过，这里选择保留以便查看
        ws_data.cell(row=current_write_row, column=1, value=keyword)
        for idx, val in enumerate(sums):
            ws_data.cell(row=current_write_row, column=idx + 2, value=val)

        valid_keywords_order.append(keyword)
        current_write_row += 1

    # 5. 画图逻辑 (复用并修改之前的逻辑)
    if 'Charts' in wb.sheetnames:
        charts_sheet = wb['Charts']
        charts_sheet.delete_rows(1, charts_sheet.max_row)
        charts_sheet.delete_cols(1, charts_sheet.max_column)
    else:
        charts_sheet = wb.create_sheet("Charts")

    chart_width = 20
    chart_height = 10
    vertical_gap = 10
    current_chart_row = 1

    # 数据源现在是 ws_data (ProcessedData)
    # 分类标签是第一行 B列到最后一列
    categories = Reference(ws_data,
                           min_col=2, max_col=data_col_end,
                           min_row=1, max_row=1)

    # 遍历刚才写入的数据行
    for i in range(len(valid_keywords_order)):
        data_row = i + 2  # 数据从第2行开始
        keyword_name = valid_keywords_order[i]

        chart = BarChart()
        chart.type = "col"
        chart.style = 10  # 换个样式，4-10都不错
        chart.title = keyword_name
        chart.width = chart_width
        chart.height = chart_height

        # 设置数据范围
        data = Reference(ws_data,
                         min_col=2, max_col=data_col_end,
                         min_row=data_row, max_row=data_row)

        series = openpyxl.chart.Series(data, title="搜索量")
        chart.series.append(series)
        chart.set_categories(categories)

        chart.x_axis.title = "月份"
        chart.y_axis.title = "搜索量汇总"
        chart.y_axis.scaling.min = 0

        # 定位
        anchor = f"A{current_chart_row}"
        charts_sheet.add_chart(chart, anchor)

        # 在L列写入名称
        charts_sheet[f"L{current_chart_row}"] = '@' + keyword_name

        current_chart_row += chart_height + vertical_gap

    # 可选：删除其他无关Sheet，只保留这三个
    # for sheet_name in wb.sheetnames:
    #     if sheet_name not in ['Sheet1', 'Sheet2', 'Charts', 'ProcessedData']:
    #         del wb[sheet_name]

    wb.save(filename)
    print(f"处理完成，汇总数据在 'ProcessedData'，图表在 'Charts'。文件已保存。")


# 使用示例
output_file = r"C:\Users\Administrator\Desktop\工作簿9.xlsx"
# 请确保你的Excel文件里真的有Sheet2且A列有词，Sheet1格式正确
process_and_chart_with_filter(output_file)